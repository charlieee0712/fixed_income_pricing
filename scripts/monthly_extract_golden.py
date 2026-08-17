"""Extract the Monthly golden table to a flat CSV (values only, faithful).

Reads `data/Project Pricing Fixed Income Instruments.xlsm` sheet Monthly rows
100..2896 (header = row 99) and writes outputs/monthly_golden_rows.csv.
Cell values are passed through untouched: numbers as full-precision repr,
dates as ISO, error strings ("#VALUE!", "#NAME?", ...) verbatim — the recon
driver owns all parsing/classification.  Run on 47:
    PYTHONPATH=src python3 scripts/monthly_extract_golden.py
"""

import csv
import datetime as dt
import os

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

PATH = os.path.join(os.environ.get("FIP_DATA_DIR", "data"),
                    "Project Pricing Fixed Income Instruments.xlsm")
OUT = os.path.join(os.environ.get("FIP_OUT", "outputs"), "monthly_golden_rows.csv")

COLMAP = [
    ("A", "name"), ("M", "sec_id"), ("B", "price"), ("C", "cpn"),
    ("D", "cpn_typ"), ("E", "cpn_freq"), ("F", "ccy"), ("G", "maturity"),
    ("H", "day_cnt"), ("I", "bbg_eff_dur"), ("J", "rating"), ("K", "par_held"),
    ("N", "vol"), ("O", "valuation"), ("P", "P_oas"), ("Q", "Q_dur"),
    ("R", "R_widen"), ("S", "S_tighten"), ("T", "T_steep"), ("U", "U_flat"),
    ("V", "V_vol_up"), ("W", "W_vol_base"), ("X", "X_base_price"),
    ("Y", "Y_vega_bp"), ("Z", "Z_vega_usd"), ("AB", "mty_typ"),
    ("AC", "calc_typ_des"), ("AD", "AD_bbg_oas"), ("AE", "AE_sys"),
    ("AG", "AG_diff"), ("AJ", "security_typ"), ("AK", "issuer"),
    ("AN", "asset_sub_class"), ("AQ", "mkt_sector"),
]


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float):
        return repr(v)
    return str(v)


def main():
    wb = load_workbook(PATH, read_only=True, data_only=True)
    ws = wb["Monthly"]
    idx = [column_index_from_string(letter) for letter, _ in COLMAP]
    maxc = max(idx)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    n = 0
    with open(OUT, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sheet_row"] + [name for _, name in COLMAP])
        for r, row in enumerate(
            ws.iter_rows(min_row=100, max_row=2896, max_col=maxc, values_only=True),
            start=100,
        ):
            vals = [cell_str(row[j - 1]) if j - 1 < len(row) else "" for j in idx]
            if not vals[0] and not vals[1]:
                continue  # fully empty row
            w.writerow([r] + vals)
            n += 1
    print(f"wrote {OUT}: {n} rows")


if __name__ == "__main__":
    main()
